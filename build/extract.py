#!/usr/bin/env python3
"""
extract.py — build-time data extractor for the PHOS dashboard.

Reads the verified financial model (data.xlsx) and emits scripts/data.js, an ES
module the static dashboard imports directly (no runtime fetch, no CDN, no xlsx
parsing in the browser). Numeric series are pulled straight from the workbook by
row label so the published figures always match the source model; qualitative
content (thesis, catalysts, risk matrix, governance) is curated alongside.

    python3 build/extract.py        # regenerate scripts/data.js

The numbers trace to SEDAR+ filings: FY2024 & FY2025 audited annuals and Q1–Q3
FY2026 interims + MDAs. See README for the full methodology.
"""
import json
import os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "data.xlsx")
OUT = os.path.join(ROOT, "scripts", "data.js")

wb = load_workbook(XLSX, data_only=True)


def sheet(name):
    return wb[name]


def find_row(ws, label, *, start=1, end=None, contains=False):
    """Return the 1-indexed row whose column-A text matches `label`."""
    end = end or ws.max_row
    for r in range(start, end + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        if (label in s) if contains else (s == label):
            return r
    raise KeyError(f"row not found: {label!r} in {ws.title!r}")


def series(ws, label, cols, *, contains=False, start=1):
    """Pull numeric values across `cols` (1-indexed) from the row for `label`."""
    r = find_row(ws, label, contains=contains, start=start)
    out = []
    for c in cols:
        v = ws.cell(row=r, column=c).value
        out.append(round(v, 6) if isinstance(v, float) else v)
    return out


# ---------------------------------------------------------------- Financials
fs = sheet("Financial Statements")
P = [2, 3, 4, 5, 6]  # B..F → FY2024, FY2025, Q1, Q2, Q3 FY2026
periods = ["FY2024", "FY2025", "Q1 FY26", "Q2 FY26", "Q3 FY26"]
period_dates = ["Feb 2024", "Feb 2025", "May 2025", "Aug 2025", "Nov 2025"]

financials = {
    "periods": periods,
    "periodDates": period_dates,
    "cash": series(fs, "Cash and cash equivalents", P, contains=True),
    "currentAssets": series(fs, "Total Current Assets", P),
    "totalAssets": series(fs, "TOTAL ASSETS", P),
    "totalLiabilities": series(fs, "TOTAL LIABILITIES", P),
    "equity": series(fs, "TOTAL SHAREHOLDERS' EQUITY", P),
    "accumDeficit": series(fs, "Accumulated deficit", P, contains=True),
    "sharesOutstanding": series(fs, "Shares Outstanding", P),
    "netLoss": series(fs, "NET LOSS AND COMPREHENSIVE LOSS", P),
    "totalExpenses": series(fs, "Total Expenses", P),
    "bookValuePerShare": series(fs, "Book Value per Share", P),
    "currentRatio": series(fs, "Current Ratio", P),
    "debtToEquity": series(fs, "Debt-to-Equity", P),
    "cashPctAssets": series(fs, "Cash as % of Total Assets", P),
}

# FY2025 full-year operating-expense breakdown (column C) for the composition chart
expense_labels = [
    ("Mining exploration & metallurgy (net)", "Exploration & metallurgy"),
    ("Share based compensation", "Share-based comp"),
    ("Business development", "Business development"),
    ("R&D / Consulting fees", "R&D / consulting"),
    ("Professional fees", "Professional fees"),
    ("General administrative", "General & admin"),
    ("Regulatory and compliance", "Regulatory"),
    ("Management / Directors' fees", "Management fees"),
]
expense_breakdown = []
for src, nice in expense_labels:
    val = series(fs, src, [3], contains=True)[0]  # FY2025 = col C
    expense_breakdown.append({"label": nice, "value": abs(val) if val else 0})
financials["expenseBreakdownFY2025"] = expense_breakdown

# ----------------------------------------------------------------- Cash burn
cb = sheet("Cash Burn Analysis")
Q = list(range(2, 10))  # B..I → 8 quarters
burn = {
    "quarters": ["Q4 FY24", "Q1 FY25", "Q2 FY25", "Q3 FY25",
                 "Q4 FY25", "Q1 FY26", "Q2 FY26", "Q3 FY26"],
    "netLoss": series(cb, "Net Loss (quarterly)", Q, contains=True),
    "cashBalance": series(cb, "Cash Balance", Q),
    "sharesOutstanding": series(cb, "Shares Outstanding", Q),
    "cumulativeDilution": [v or 0 for v in series(cb, "Cumulative Dilution", Q, contains=True)],
}

# Forward runway scenarios + capex funding gap (curated from the workbook layout)
runway = {
    "estCashJan2026": 32_600_000,
    "scenarios": [
        {"name": "Bull", "qBurn": 2_500_000, "months": 39.1, "note": "lower burn holds"},
        {"name": "Base", "qBurn": 3_200_000, "months": 30.6, "note": "current trajectory"},
        {"name": "Bear", "qBurn": 4_500_000, "months": 21.7, "note": "full drill ramp"},
    ],
}
capex_gap = {
    "items": [
        {"label": "Bégin-Lamarche mine (PEA)", "value": 675_000_000, "status": "Needs FS → financing"},
        {"label": "Phosphoric-acid facility", "value": 240_000_000, "status": "Pre-FS complete"},
        {"label": "First Saguenay pCAM plant", "value": 90_000_000, "status": "FS complete"},
    ],
    "totalRequired": 1_005_000_000,
    "currentCash": 32_600_000,
    "twoYearBurn": 21_400_000,
    "availableForCapex": 11_200_000,
    "fundingGap": 993_800_000,
}

# ------------------------------------------------------------------- Peers
# Curated from the Peer Analysis sheet (Mar 2026 market data). mktNpv = Mkt Cap / PEA-NPV.
peers = {
    "navDiscount": [
        {"ticker": "PHOS", "name": "First Phosphate", "mktNpv": 0.10, "self": True},
        {"ticker": "LCE",  "name": "Century Lithium", "mktNpv": 0.04},
        {"ticker": "DAN",  "name": "Arianne Phosphate", "mktNpv": 0.025},
        {"ticker": "NMG",  "name": "Nouveau Monde", "mktNpv": 0.34},
        {"ticker": "PMET", "name": "PMET Resources", "mktNpv": 0.37},
    ],
    "irr": [
        {"ticker": "PHOS", "name": "First Phosphate", "irr": 0.33, "self": True},
        {"ticker": "LCE",  "name": "Century Lithium", "irr": 0.274},
        {"ticker": "NMG",  "name": "Nouveau Monde", "irr": 0.20},
        {"ticker": "DAN",  "name": "Arianne Phosphate", "irr": 0.187},
        {"ticker": "PMET", "name": "PMET Resources", "irr": 0.181},
    ],
    "table": [
        {"ticker": "PHOS", "name": "First Phosphate", "component": "Phosphate → cathode",
         "stage": "PEA (2024)", "mktCap": "C$158M", "npv": "C$1,590M", "irr": "33%",
         "capex": "C$675M", "mktNpv": "0.10×", "self": True},
        {"ticker": "DAN", "name": "Arianne Phosphate", "component": "Phosphate (concentrate)",
         "stage": "FS (2013)", "mktCap": "C$55M", "npv": "C$2,200M", "irr": "18.7%",
         "capex": "C$1,200M", "mktNpv": "0.025×"},
        {"ticker": "NMG", "name": "Nouveau Monde", "component": "Graphite → anode",
         "stage": "Pre-construction", "mktCap": "US$344M", "npv": "~US$1B", "irr": "~20%",
         "capex": "~C$1B+", "mktNpv": "0.34×"},
        {"ticker": "PMET", "name": "PMET Resources", "component": "Lithium → cathode",
         "stage": "FS (2025)", "mktCap": "C$584M", "npv": "C$1,594M", "irr": "18.1%",
         "capex": "C$2,600M", "mktNpv": "0.37×"},
        {"ticker": "LCE", "name": "Century Lithium", "component": "Lithium → cathode",
         "stage": "FS (2026)", "mktCap": "C$106M", "npv": "~US$3B", "irr": "27.4%",
         "capex": "~US$1B", "mktNpv": "0.04×"},
    ],
}

# ---------------------------------------------------------------- Valuation
vm = sheet("Valuation Model")
valuation = {
    "npv": series(vm, "After-Tax NPV (8%) from PEA", [2], contains=True)[0],
    "currentPrice": series(vm, "Current Share Price", [2], contains=True)[0],
    "sharesBasic": series(vm, "Shares Outstanding (basic)", [2], contains=True)[0],
    "sharesFD": series(vm, "Fully Diluted Shares", [2], contains=True)[0],
    "sensitivityPct": series(vm, "% of NPV Applied", list(range(2, 10)), contains=True),
    "impliedBasic": series(vm, "Implied Price/Share (basic)", list(range(2, 10)), contains=True),
    "impliedFD": series(vm, "Implied Price/Share (FD)", list(range(2, 10)), contains=True),
    "scenarios": [
        {"name": "Bear", "pct": 0.03, "price": 0.275, "label": "PEA-stage, 83% Inferred",
         "rationale": "Typical PEA-stage explorer discount; speculative resource."},
        {"name": "Base", "pct": 0.11, "price": 1.009, "label": "Where PHOS trades",
         "rationale": "Current market — LFP premium already in the price.", "current": True},
        {"name": "Bull", "pct": 0.20, "price": 1.835, "label": "FS complete, de-risked",
         "rationale": "Post-FS, financing secured, clearer path to production."},
        {"name": "Moon", "pct": 0.35, "price": 3.212, "label": "Construction / FID",
         "rationale": "FID made, construction underway, major dilution factored in."},
    ],
}

# ----------------------------------------------------------- Snapshot / KPIs
snapshot = {
    "price": 1.05,
    "currency": "C$",
    "marketCap": 158_000_000,
    "marketCapFD": 202_000_000,
    "cash": 19_983_238,
    "cashEst": 32_600_000,
    "runwayMonths": 24,
    "mktNpv": 0.10,
    "netLossAnnualized": 12_700_000,
    "sharesOutstanding": 151_218_841,
    "sharesFD": 192_400_000,
    "ytdDilution": 1.049,
    "weekRange": "C$0.24 – C$1.13",
}

# ------------------------------------------------------------------- Project (PEA)
project = {
    "name": "Bégin-Lamarche",
    "npv": 1_590_000_000,
    "irr": 0.33,
    "capex": 675_000_000,
    "payback": 2.9,
    "mineLife": 23,
    "production": "1 Mtpa concentrate",
    "resource": 255.5,
    "inferredPct": 0.83,
    "grade": "6.0–8.2% P₂O₅",
    "opex": "US$144/t",
    "royalties": "None (100%-owned)",
}

# ------------------------------------------------------------------- Thesis
thesis = {
    "bull": [
        "The only pure-play LFP-phosphate company on North American exchanges — no direct comparable.",
        "Already produced commercial-grade LFP 18650 cells: proof of the mine-to-battery chain.",
        "Strong government backing — C$16.7M federal grant plus critical-mineral tax credits.",
        "PEA shows a 33% after-tax IRR and 2.9-year payback — best-in-class among junior peers.",
        "LFP reached ~50% of EV battery chemistry in 2024; phosphate is the supply-constrained node.",
    ],
    "base": [
        "Trades at ~10% of PEA NAV — a premium to phosphate juniors, a discount to LFP-chain peers.",
        "Well-funded near term (~C$33M, ~24-month runway) but pre-revenue until ~2029.",
        "Re-rating hinges on near-term catalysts: drill results → PFS → financing.",
        "LFP optionality is the whole premium — fertilizer economics are the floor.",
    ],
    "bear": [
        "83% of the resource is Inferred — speculative until a feasibility study confirms it.",
        "A C$675M+ initial-capex gap with no committed financing path.",
        "105%+ share dilution since FY2024, with more required to fund development.",
        "Going-concern audit opinion; no revenue for years.",
        "PEA assumptions may not survive a feasibility study.",
    ],
}

# ------------------------------------------------------------------ Catalysts
catalysts = [
    {"date": "Mar 2026", "event": "C$16.7M non-repayable federal contribution",
     "weight": "major", "note": "Government validation; narrows the capex gap."},
    {"date": "Feb 2026", "event": "Phosphate added to Canada's critical-minerals list",
     "weight": "major", "note": "Unlocks 30% CMETC + critical-mineral tax credits."},
    {"date": "Feb 2026", "event": "ADR program launched (OTCQX: FPHOY)",
     "weight": "minor", "note": "Broader US institutional access."},
    {"date": "Jan 2026", "event": "Initial offtake-agreement payment received",
     "weight": "minor", "note": "First commercial-revenue milestone."},
    {"date": "Dec 2025", "event": "Added to the CSE25 Index",
     "weight": "minor", "note": "Passive-fund visibility."},
    {"date": "Dec 2025", "event": "C$9.6M financing closed",
     "weight": "minor", "note": "Extended runway."},
    {"date": "Jul 2025", "event": "Commercial-grade LFP 18650 cells produced",
     "weight": "major", "note": "Mine-to-battery proof of concept."},
]
forward_catalysts = [
    {"window": "Q1–Q2 2026", "event": "30,000 m drill-program results",
     "impact": "Resource upgrade: Inferred → Indicated"},
    {"window": "2026", "event": "Pre-Feasibility Study initiation",
     "impact": "Major de-risking event"},
    {"window": "2026–27", "event": "Strategic JV / partnership for capex",
     "impact": "De-risks the funding gap"},
    {"window": "2027+", "event": "Feasibility Study completion",
     "impact": "Unlocks project finance; NAV re-rate"},
]

# ----------------------------------------------------------------- Risk matrix
SEV = {"Low": 1, "Medium": 2, "High": 3, "Critical": 4}
PROB = {"Low": 1, "Medium": 2, "High": 3}
risks = [
    {"factor": "Capex funding failure", "severity": "Critical", "probability": "Medium",
     "mitigation": "Govt grants, JV potential"},
    {"factor": "Continued dilution", "severity": "High", "probability": "High",
     "mitigation": "None — inherent to the model"},
    {"factor": "Resource downgrade in PFS/FS", "severity": "High", "probability": "Medium",
     "mitigation": "30,000 m drill program underway"},
    {"factor": "Permitting delays", "severity": "Medium", "probability": "Medium",
     "mitigation": "QC mining-friendly; critical mineral"},
    {"factor": "LFP chemistry shift", "severity": "Medium", "probability": "Low",
     "mitigation": "LFP gaining share; fertilizer fallback"},
]
for r in risks:
    r["sev"] = SEV[r["severity"]]
    r["prob"] = PROB[r["probability"]]

# ----------------------------------------------------------------- Governance
governance = {
    "ratings": [
        {"area": "Technical expertise", "stars": 5},
        {"area": "Strategic relationships", "stars": 5},
        {"area": "Compensation alignment", "stars": 5},
        {"area": "Insider conviction", "stars": 5},
        {"area": "Executive leadership", "stars": 4},
        {"area": "Board independence", "stars": 3},
        {"area": "Operational readiness", "stars": 3},
    ],
    "overall": 4,
    "highlights": [
        {"k": "CEO compensation", "v": "100% equity, C$0 cash",
         "note": "Maximum alignment — the CEO only wins if the share price does."},
        {"k": "CEO open-market buying", "v": "~C$1.8M",
         "note": "2.87M shares bought with personal capital since May 2023."},
        {"k": "CEO ownership", "v": "~17.3%",
         "note": "Largest individual shareholder; also a key-man risk."},
        {"k": "Board alignment", "v": "100% RSU fees",
         "note": "Directors paid entirely in equity since Sep 2023."},
    ],
    "insiders": [
        {"name": "John Passalacqua", "role": "CEO & Director", "bought": "2,872,000 sh (~C$1.8M)"},
        {"name": "Peter Nicholson", "role": "Independent Director", "bought": "532,000 sh (C$146K)"},
        {"name": "Laurence Zeifman", "role": "Chairman", "bought": "359,500 sh (C$74K)"},
    ],
}

# ------------------------------------------------------- Project deep-dive (PEA)
# Bégin-Lamarche PEA (2024). Figures from the workbook's project & peer economics.
project_detail = {
    "summary": "Bégin-Lamarche is an igneous (anorthosite-hosted) phosphate project in the "
               "Saguenay–Lac-St-Jean region of Québec. Its 2024 PEA outlines a 23-year, "
               "1 Mtpa concentrate operation feeding a downstream purified-phosphoric-acid and "
               "LFP-cathode strategy — the basis of the mine-to-battery thesis.",
    "economics": [
        {"k": "After-tax NPV (8%)", "v": "C$1,590M"},
        {"k": "After-tax IRR", "v": "33%"},
        {"k": "Initial capex", "v": "C$675M"},
        {"k": "Payback", "v": "2.9 years"},
        {"k": "Mine life", "v": "23 years"},
        {"k": "Production", "v": "1 Mtpa concentrate"},
        {"k": "Operating cost", "v": "US$144 / t conc."},
        {"k": "Revenue at production", "v": "~C$600M / yr"},
        {"k": "Royalties", "v": "None (100%-owned)"},
        {"k": "Land position", "v": "~1,500 km² claims"},
    ],
    # Resource classification — 255.5 Mt, 83% Inferred (the key geological risk)
    "resource": {
        "total": 255.5, "grade": "6.0–8.2% P₂O₅",
        "split": [
            {"label": "Inferred", "pct": 0.83, "note": "speculative — needs FS"},
            {"label": "Measured + Indicated", "pct": 0.17, "note": "higher confidence"},
        ],
    },
    # Downstream / vertical integration steps (mine → battery)
    "downstream": [
        {"step": "Mine & concentrator", "detail": "Igneous phosphate → high-purity concentrate", "capex": "C$675M", "status": "PEA"},
        {"step": "Purified phosphoric acid", "detail": "Battery-grade PPA (low-contaminant)", "capex": "C$240M", "status": "Pre-FS"},
        {"step": "pCAM / LFP cathode", "detail": "First Saguenay cathode-active-material plant", "capex": "C$90M", "status": "FS complete"},
        {"step": "LFP cells", "detail": "Commercial-grade 18650 cells demonstrated (Jul 2025)", "capex": "—", "status": "Proven"},
    ],
}

# ------------------------------------------------------------- Market & industry
market = {
    "lfpShareEV": 0.50,          # LFP ≈ 50% of EV battery chemistry in 2024 (up from 37% in 2022)
    "lfpShare2022": 0.37,
    "chinaLfpPowder": 0.90,      # China controls ~90% of LFP powder production
    "phosphatePrice": "~US$600 / t (DAP/MAP)",
    "marketGrowth": {            # LFP battery market size, US$B
        "years": ["2024", "2027", "2030", "2032"],
        "low": [17, 30, 48, 64],
        "high": [17, 38, 62, 84],
        "cagr": "17–19%",
    },
    "drivers": [
        {"k": "EV adoption", "v": "LFP overtook NMC to ~50% of EV chemistry in 2024", "note": "Cost & safety favour LFP for mass-market EVs."},
        {"k": "Grid storage (ESS)", "v": "ESS demand +50% YoY (J.P. Morgan)", "note": "LFP dominates stationary storage."},
        {"k": "Supply concentration", "v": "China ~90% of LFP powder", "note": "Western supply chains seek non-China phosphate."},
        {"k": "Critical-mineral policy", "v": "Phosphate listed US (Nov 2025) & Canada (Feb 2026)", "note": "Unlocks tax credits & fast-track permitting."},
    ],
}

# ------------------------------------------------------- Supply-chain positioning
supply_chain = {
    "chain": ["Phosphate rock", "Concentrate", "Purified phosphoric acid", "pCAM", "LFP cathode", "LFP cell"],
    "phosNode": 2,   # PHOS spans rock → PPA → (planned) pCAM/cathode
    "cluster": [
        {"ticker": "PHOS", "input": "Phosphate", "self": True},
        {"ticker": "NMG", "input": "Graphite (anode)"},
        {"ticker": "PMET", "input": "Lithium (cathode)"},
    ],
    "insight": "Three of the four North-American LFP-input juniors are Québec-based (PHOS phosphate, "
               "NMG graphite, PMET lithium), creating a natural cluster for government support and logistics. "
               "PHOS has already produced commercial LFP 18650 cells using its own phosphate and NMG graphite — "
               "the only junior in the group to demonstrate the full mine-to-battery chain.",
}

# ----------------------------------------------------- Capital structure & owners
capital = {
    "sharesBasic": 151_218_841,
    "sharesCurrent": 173_267_217,
    "sharesFD": 192_400_000,
    "omnibusReservedPct": 0.096,     # 14.58M reserved ≈ 9.6% of o/s
    "omnibusCapPct": 0.20,           # rolling 20% cap
    "debt": "Debt-free (FT premium only)",
    "financingMethod": "Flow-through + placements",
    "ownership": [
        {"name": "John Passalacqua (CEO)", "pct": 0.173},
        {"name": "L. Zeifman (Chairman)", "pct": 0.061},
        {"name": "B. Kurtz (CFO)", "pct": 0.054},
        {"name": "P. Nicholson (Director)", "pct": 0.028},
        {"name": "Other insiders / float", "pct": 0.684},
    ],
}

# --------------------------------------------------------- Financing & gov support
financing = [
    {"date": "Mar 2026", "item": "C$16.7M non-repayable federal contribution", "type": "Non-dilutive", "note": "Conditional; narrows the capex gap."},
    {"date": "Feb 2026", "item": "Critical-mineral designation (Canada)", "type": "Tax credits", "note": "30% CMETC + critical-mineral tax credits."},
    {"date": "Dec 2025", "item": "C$9.6M financing closed", "type": "Equity", "note": "Extended runway to mid-2027."},
    {"date": "FY2026", "item": "Flow-through share placements", "type": "Equity (tax-advantaged)", "note": "Primary funding method; drives dilution."},
    {"date": "Future", "item": "Project finance / JV for capex", "type": "Debt / strategic", "note": "Required to close the ~C$994M gap."},
]

# --------------------------------------------------- ESG · permitting · Indigenous
esg = [
    {"area": "Indigenous relations", "rating": 5, "detail": "President Armand MacKenzie helped draft the UN Declaration on the Rights of Indigenous Peoples; leading a partnership with the Pekuakamiulnuatsh First Nation for deep Indigenous financial participation."},
    {"area": "Permitting jurisdiction", "rating": 4, "detail": "Québec is consistently ranked among the most mining-friendly jurisdictions; phosphate's critical-mineral status supports fast-tracking."},
    {"area": "Resource ownership", "rating": 5, "detail": "100%-owned claims (~1,500 km²) with no royalties — clean title and full upside retention."},
    {"area": "Community / local roots", "rating": 4, "detail": "SVP David Dufour brings 30+ years of Saguenay operating and community experience; local social licence is a strategic priority."},
    {"area": "Governance independence", "rating": 3, "detail": "Only 2 of 5 directors are independent — below TSXV best practice; a key-man dependency on the CEO."},
]

# ---------------------------------------------------- Valuation comps & football field
comps = {
    "pnav": [   # market cap / study NAV (×) — lower = cheaper
        {"ticker": "PHOS", "v": 0.10, "self": True},
        {"ticker": "DAN", "v": 0.025}, {"ticker": "LCE", "v": 0.04},
        {"ticker": "NMG", "v": 0.34}, {"ticker": "PMET", "v": 0.37},
    ],
    "evResource": [  # EV per tonne of resource (C$/t)
        {"ticker": "PHOS", "v": 0.62, "self": True},
        {"ticker": "DAN", "v": 0.09},
    ],
    # Fair-value ranges by method (C$/share) for the football field
    "footballField": [
        {"method": "PEA-NAV (bear–bull)", "low": 0.28, "high": 1.83, "mid": 1.01,
         "note": "3% → 20% of after-tax NPV"},
        {"method": "Peer P/NAV re-rate", "low": 0.92, "high": 2.29, "mid": 1.38,
         "note": "to junior-peer 10–25% of NAV"},
        {"method": "EV / resource (vs DAN)", "low": 0.45, "high": 1.38, "mid": 0.92,
         "note": "5–15% of NPV per tonne"},
        {"method": "Upside (post-FS / FID)", "low": 1.83, "high": 3.21, "mid": 2.52,
         "note": "de-risked → construction", "upside": True},
    ],
    "fairValueLow": 0.28,
    "fairValueHigh": 1.83,
}

# ------------------------------------------------------------ Roadmap to production
roadmap = [
    {"phase": "PEA", "when": "2024", "done": True, "detail": "NPV C$1.59B · IRR 33% established"},
    {"phase": "30,000 m drill program", "when": "2026", "done": False, "detail": "Upgrade Inferred → Indicated resource"},
    {"phase": "Pre-Feasibility Study", "when": "2026", "done": False, "detail": "First major de-risking"},
    {"phase": "Feasibility Study", "when": "2027", "done": False, "detail": "Bankable economics → project finance"},
    {"phase": "Financing / FID", "when": "2027–28", "done": False, "detail": "Debt · JV · strategic · equity"},
    {"phase": "Construction", "when": "2028", "done": False, "detail": "Mine + downstream build-out"},
    {"phase": "First production", "when": "~2029", "done": False, "detail": "Concentrate → revenue"},
]

# --------------------------------------------------------------- Advisory board
advisory = [
    {"name": "Yves Caprara", "role": "Phosphoric-acid technology", "stars": 5,
     "note": "Former CEO of Prayon SA — Europe's largest food/technical phosphoric-acid producer. The single most relevant advisor to the PPA thesis."},
    {"name": "Bernard Lapointe, PhD", "role": "Québec phosphate", "stars": 5,
     "note": "Founded Arianne Phosphate (the only other significant Québec phosphate company); proved phosphate viability in Saguenay–Lac-St-Jean."},
    {"name": "Gary Stanley", "role": "US critical-minerals policy", "stars": 5,
     "note": "Former Director, Office of Critical Minerals & Metals, US Dept. of Commerce; lead author of the 2019 US Critical Minerals Strategy."},
    {"name": "Mario Bouchard", "role": "Québec government finance", "stars": 5,
     "note": "Former Assistant Deputy Minister for Strategic Industries — insider knowledge of provincial project financing, critical for the capex."},
    {"name": "Dr. Peir Pufahl, P.Geo", "role": "Phosphorite geology", "stars": 4,
     "note": "Queen's University professor; internationally recognised phosphorite expert partnered with PHOS on its mineralogy."},
]

# ------------------------------------------------------------------- Glossary
glossary = [
    {"t": "PEA", "d": "Preliminary Economic Assessment — the earliest, least-certain economic study of a mining project."},
    {"t": "PFS / FS", "d": "Pre-Feasibility / Feasibility Study — progressively more rigorous, bankable studies."},
    {"t": "NI 43-101", "d": "Canadian standard for disclosing mineral resources/reserves; classes are Inferred, Indicated, Measured."},
    {"t": "Inferred resource", "d": "Lowest-confidence resource class — cannot be used in a feasibility study or reserves."},
    {"t": "NAV / NPV", "d": "Net asset value / net present value — the discounted value of a project's future cash flows."},
    {"t": "IRR", "d": "Internal rate of return — the discount rate at which a project's NPV is zero."},
    {"t": "LFP", "d": "Lithium iron phosphate — a low-cost, safe battery cathode chemistry (Li-Fe-PO₄)."},
    {"t": "PPA", "d": "Purified phosphoric acid — battery-grade input for cathode-active material."},
    {"t": "pCAM / CAM", "d": "(precursor) cathode-active material — the processed feed for battery cathodes."},
    {"t": "Flow-through shares", "d": "A Canadian tax-advantaged financing where exploration expenses pass through to investors."},
    {"t": "CMETC", "d": "Critical Mineral Exploration Tax Credit — a 30% Canadian credit for eligible critical-mineral work."},
    {"t": "P₂O₅", "d": "Phosphorus pentoxide — the standard measure of phosphate grade/content."},
]

# --------------------------------------------------------------- Coverage box
coverage = {
    "assessment": "High-risk / high-reward — speculative",
    "fairValueLow": 0.28, "fairValueHigh": 1.83,
    "staticPrice": 1.05,
    "stats": [
        {"k": "Mkt cap", "v": "C$158M"},
        {"k": "Mkt cap / NAV", "v": "~10%"},
        {"k": "PEA NPV", "v": "C$1.59B"},
        {"k": "PEA IRR", "v": "33%"},
        {"k": "Cash runway", "v": "~24 mo"},
        {"k": "Stage", "v": "PEA · pre-revenue"},
    ],
}

# --------------------------------------------------------------- Monitoring
monitoring = [
    {"metric": "Quarterly cash balance", "current": "C$20.0M (Q3)", "watch": "Below C$10M"},
    {"metric": "Shares outstanding", "current": "~173M", "watch": "Above 200M"},
    {"metric": "Mkt cap / NAV", "current": "~10%", "watch": ">15% rich · <5% distress"},
    {"metric": "Resource classification", "current": "83% Inferred", "watch": "Upgrade → Indicated"},
    {"metric": "LFP share of EV batteries", "current": "~50%", "watch": "Shift back to NMC"},
    {"metric": "Phosphate price (DAP/MAP)", "current": "~US$600/t", "watch": "Rising = positive"},
]

# --------------------------------------------------------------- Data vintage
vintage = {
    "filings": "SEDAR+ filings through Q3 FY2026 (Nov 30, 2025)",
    "market": "Peer & market data as of March 2026",
    "price": "Share price live (15-min delayed) when available; else C$1.05 (Mar 2026)",
}

# --------------------------------------------------------------------- Bundle
DATA = {
    "company": {
        "name": "First Phosphate Corp.",
        "tickers": ["CSE: PHOS", "OTCQX: FRSPF", "FSE: KD0"],
        "sector": "Critical minerals · LFP battery supply chain",
        "hq": "Saguenay, QC · Vancouver, BC",
        "fyEnd": "February 28/29",
        "stage": "Pre-revenue · exploration–development",
    },
    "asOf": "March 2026",
    "snapshot": snapshot,
    "project": project,
    "financials": financials,
    "burn": burn,
    "runway": runway,
    "capexGap": capex_gap,
    "peers": peers,
    "valuation": valuation,
    "thesis": thesis,
    "catalysts": catalysts,
    "forwardCatalysts": forward_catalysts,
    "risks": risks,
    "governance": governance,
    "projectDetail": project_detail,
    "market": market,
    "supplyChain": supply_chain,
    "capital": capital,
    "financing": financing,
    "esg": esg,
    "comps": comps,
    "roadmap": roadmap,
    "advisory": advisory,
    "glossary": glossary,
    "coverage": coverage,
    "monitoring": monitoring,
    "vintage": vintage,
    "sources": [
        "FY2024 & FY2025 audited annual financial statements (SEDAR+)",
        "Q1–Q3 FY2026 interim financial statements (SEDAR+)",
        "Management Discussion & Analysis, Q1–Q3 FY2026",
        "Peer & market data from public sources (Mar 2026)",
    ],
}

banner = (
    "// AUTO-GENERATED by build/extract.py from data.xlsx — do not edit by hand.\n"
    "// Figures trace to SEDAR+ filings (FY2024–Q3 FY2026). Regenerate: python3 build/extract.py\n"
)
with open(OUT, "w") as f:
    f.write(banner)
    f.write("export const DATA = ")
    f.write(json.dumps(DATA, indent=2, ensure_ascii=False))
    f.write(";\nexport default DATA;\n")

print(f"wrote {OUT}")
print(f"  periods: {financials['periods']}")
print(f"  cash:    {financials['cash']}")
print(f"  netLoss: {financials['netLoss']}")
print(f"  NPV:     {valuation['npv']:,}  price: {valuation['currentPrice']}")
print(f"  sensitivity pct:    {valuation['sensitivityPct']}")
print(f"  implied (basic):    {valuation['impliedBasic']}")
